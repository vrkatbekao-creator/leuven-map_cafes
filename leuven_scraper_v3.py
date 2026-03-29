"""
Leuven Tourism × Cafe — Google Places 整合爬蟲 v3
===================================================
v3 新增：景點多語系名稱與描述（EN / NL / FR / DE / ZH-TW）
GeoJSON 每個景點 Feature 包含：
  properties.names        = {en, nl, fr, de, zh}
  properties.descriptions = {en, nl, fr, de, zh}
地圖依使用者選擇的語言顯示對應文字。

使用:
  python leuven_scraper_v3.py --google-key YOUR_KEY
  python leuven_scraper_v3.py --google-key YOUR_KEY --radius 3000
"""

import argparse, json, math, os, re, time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LEUVEN_LAT = 50.8798
LEUVEN_LON = 4.7005
OUTPUT_DIR = "leuven_output"
EXCEL_FILE = "leuven_tourism_cafes.xlsx"
JSON_FILE  = "leuven_data.json"
THEME      = "1A6B3C"

WIFI_KW   = ["wifi","wi-fi","wireless","internet","hotspot"]
SOCKET_KW = ["socket","outlet","plug","charging","power point",
             "power outlet","electrical","laptop friendly","cowork"]
QUIET_KW  = ["quiet","calm","peaceful","good for work","study",
             "concentration","not noisy","tranquil"]

# ─────────────────────────────────────────────────────────────
# 多語系景點資料
# 每筆: (canonical_name, category, lat, lon, slug,
#        {en, nl, fr, de, zh} names,
#        {en, nl, fr, de, zh} descriptions)
# ─────────────────────────────────────────────────────────────
ATTRACTIONS_ML = [
  (
    "Stadhuis & Grote Markt", "highlight", 50.8799, 4.7006, "stadhuis",
    dict(en="Stadhuis & Grote Markt",
         nl="Stadhuis & Grote Markt",
         fr="Hôtel de ville & Grande Place",
         de="Rathaus & Großer Markt",
         zh="市政廳與大廣場"),
    dict(en="One of Belgium's finest Gothic town halls, overlooking the vibrant main square.",
         nl="Een van de mooiste gotische stadhizen van België, aan het bruisende Grote Markt.",
         fr="L'un des plus beaux hôtels de ville gothiques de Belgique, sur la place centrale.",
         de="Eines der schönsten gotischen Rathäuser Belgiens am lebhaften Hauptplatz.",
         zh="比利時最精美的哥德式市政廳之一，坐落於熱鬧的市中心廣場。"),
  ),
  (
    "Oude Markt", "highlight", 50.8790, 4.7038, "oude-markt",
    dict(en="Oude Markt",nl="Oude Markt",fr="Vieux Marché",de="Alter Markt",zh="舊市場廣場"),
    dict(en="Europe's longest bar: over 40 cafes and bars packed into one lively square.",
         nl="Europa's langste toog: meer dan 40 cafés en bars op één bruisend plein.",
         fr="Le plus long bar d'Europe : plus de 40 cafés et bars sur une seule place animée.",
         de="Europas längste Theke: über 40 Cafés und Bars auf einem einzigen lebhaften Platz.",
         zh="歐洲最長的酒吧街：超過40間咖啡館與酒吧匯聚於一個熱鬧廣場。"),
  ),
  (
    "Vaartkom", "highlight", 50.8857, 4.7082, "vaartkom",
    dict(en="Vaartkom",nl="Vaartkom",fr="Vaartkom",de="Vaartkom",zh="運河區 Vaartkom"),
    dict(en="Former industrial canal district transformed into a hip hub of restaurants and bars.",
         nl="Voormalige industriële kanaalwijk, nu een trendy hotspot met restaurants en bars.",
         fr="Ancien quartier industriel au bord du canal, devenu un lieu branché de restaurants.",
         de="Ehemaliges Industriekanalviertel, heute trendiger Hotspot mit Restaurants und Bars.",
         zh="前工業運河區，如今搖身一變為充滿活力的餐廳與酒吧聚集地。"),
  ),
  (
    "M Leuven", "museum", 50.8773, 4.6991, "mleuven",
    dict(en="M Leuven",nl="M Leuven",fr="M Leuven",de="M Leuven",zh="魯汶市立博物館 M"),
    dict(en="Leuven's city museum spanning art and history, including Roger van der Weyden masterworks.",
         nl="Het stadsmuseum van Leuven met kunst en geschiedenis, inclusief werken van Van der Weyden.",
         fr="Le musée municipal de Louvain, de l'art et de l'histoire, avec des œuvres de Van der Weyden.",
         de="Leuven's Stadtmuseum mit Kunst und Geschichte, u.a. Werke von Roger van der Weyden.",
         zh="魯汶市立博物館，館藏橫跨藝術與歷史，收有羅希爾·范德魏登名作。"),
  ),
  (
    "Sint-Pieterskerk", "museum", 50.8792, 4.7009, "sint-pieterskerk",
    dict(en="Sint-Pieterskerk",nl="Sint-Pieterskerk",fr="Église Saint-Pierre",
         de="St.-Peterskirche",zh="聖彼得教堂"),
    dict(en="Gothic church on the Grote Markt, housing Dieric Bouts' acclaimed triptych altarpiece.",
         nl="Gotische kerk aan de Grote Markt met het beroemde drieluik van Dieric Bouts.",
         fr="Église gothique abritant le célèbre triptyque de Thierry Bouts.",
         de="Gotische Kirche mit dem berühmten Triptychon von Dieric Bouts.",
         zh="坐落於大廣場的哥德式教堂，珍藏迪爾克·鮑茨的三聯祭壇畫傑作。"),
  ),
  (
    "PARCUM", "museum", 50.8752, 4.6935, "parcum",
    dict(en="PARCUM",nl="PARCUM",fr="PARCUM",de="PARCUM",zh="宗教藝術博物館 PARCUM"),
    dict(en="Museum of religious art and culture with an exceptional collection of sacred objects.",
         nl="Museum voor religieuze kunst en cultuur met een uitzonderlijke collectie sacrale objecten.",
         fr="Musée d'art et de culture religieux avec une collection exceptionnelle d'objets sacrés.",
         de="Museum für religiöse Kunst und Kultur mit einer außergewöhnlichen Sammlung sakraler Objekte.",
         zh="宗教藝術與文化博物館，典藏豐富的宗教聖物與藝術珍品。"),
  ),
  (
    "Groot Begijnhof", "religious", 50.8736, 4.6971, "groot-begijnhof",
    dict(en="Groot Begijnhof",nl="Groot Begijnhof",fr="Grand Béguinage",
         de="Großes Beginenhof",zh="大貝居安院"),
    dict(en="UNESCO World Heritage beguinage — a perfectly preserved medieval walled village within the city.",
         nl="UNESCO Werelderfgoed begijnhof — een perfect bewaard middeleeuws dorp binnen de stad.",
         fr="Béguinage classé au patrimoine mondial de l'UNESCO — village médiéval intact au cœur de la ville.",
         de="UNESCO-Weltkulturerbe — ein perfekt erhaltenes mittelalterliches Dorf in der Stadt.",
         zh="聯合國教科文組織世界遺產，城市中保存完好的中世紀圍牆村落。"),
  ),
  (
    "Klein Begijnhof", "religious", 50.8825, 4.7082, "klein-begijnhof",
    dict(en="Klein Begijnhof",nl="Klein Begijnhof",fr="Petit Béguinage",
         de="Kleines Beginenhof",zh="小貝居安院"),
    dict(en="A smaller, quieter beguinage north of the city centre, less visited and more tranquil.",
         nl="Een kleiner, rustiger begijnhof ten noorden van het stadscentrum, minder bezocht.",
         fr="Un béguinage plus petit et tranquille au nord du centre-ville, moins fréquenté.",
         de="Ein kleinerer, ruhigerer Beginenhof nördlich des Stadtzentrums, weniger besucht.",
         zh="位於市中心北側的小型貝居安院，訪客較少，環境清幽。"),
  ),
  (
    "Abdij van Park", "religious", 50.8506, 4.6895, "abdij-van-park-0",
    dict(en="Abdij van Park",nl="Abdij van Park",fr="Abbaye du Parc",
         de="Abtei von Park",zh="帕克修道院"),
    dict(en="A 12th-century Norbertine abbey with a working farm, brewery, and stunning gardens.",
         nl="Een 12e-eeuwse norbertijnerabdij met een werkende hoeve, brouwerij en prachtige tuinen.",
         fr="Abbaye norbertine du XIIe siècle avec ferme active, brasserie et magnifiques jardins.",
         de="Eine Norbertiner-Abtei aus dem 12. Jahrhundert mit Bauernhof, Brauerei und Gärten.",
         zh="12世紀諾伯特修道院，設有運作中的農場、釀酒廠與精美庭園。"),
  ),
  (
    "Abdij Keizersberg", "religious", 50.8830, 4.6940, "abdij-keizersberg",
    dict(en="Abdij Keizersberg",nl="Abdij Keizersberg",fr="Abbaye du Mont César",
         de="Abtei Keizersberg",zh="凱澤斯貝格修道院"),
    dict(en="Benedictine abbey perched on a hill with peaceful gardens overlooking the city.",
         nl="Benedictijnenabdij op een heuvel met rustige tuinen en uitzicht over de stad.",
         fr="Abbaye bénédictine sur une colline avec des jardins paisibles dominant la ville.",
         de="Benediktinerabtei auf einem Hügel mit ruhigen Gärten und Blick über die Stadt.",
         zh="本篤會修道院，建於山丘之上，庭園清幽，可俯瞰整座城市。"),
  ),
  (
    "Universiteitsbibliotheek", "university", 50.8779, 4.7024, "universiteitsbibliotheek-toren",
    dict(en="University Library & Tower",nl="Universiteitsbibliotheek & -toren",
         fr="Bibliothèque universitaire & tour",de="Universitätsbibliothek & Turm",
         zh="魯汶大學圖書館與鐘樓"),
    dict(en="KU Leuven's iconic library tower, rebuilt after WWI destruction with donations from the USA.",
         nl="De iconische bibliotheektoren van KU Leuven, herbouwd na WO I met Amerikaanse giften.",
         fr="La tour emblématique de la bibliothèque de KU Leuven, reconstruite après la Première Guerre mondiale.",
         de="Der ikonische Bibliotheksturm der KU Leuven, nach dem Ersten Weltkrieg mit US-Spenden wiederaufgebaut.",
         zh="魯汶大學標誌性圖書館塔樓，一戰後由美國捐款重建，是城市的精神象徵。"),
  ),
  (
    "Universiteitshal", "university", 50.8794, 4.7018, "universiteitshal",
    dict(en="Universiteitshal",nl="Universiteitshal",fr="Halles universitaires",
         de="Universitätshalle",zh="大學禮堂"),
    dict(en="The grand ceremonial hall of KU Leuven, a symbol of the university's centuries of history.",
         nl="De grootse ceremoniezaal van KU Leuven, symbool van de eeuwenlange universiteitsgeschiedenis.",
         fr="La grande salle cérémonielle de KU Leuven, symbole des siècles d'histoire universitaire.",
         de="Die feierliche Haupthalle der KU Leuven, Symbol der jahrhundertelangen Universitätsgeschichte.",
         zh="魯汶大學莊嚴的典禮大廳，見證了大學數百年的歷史傳承。"),
  ),
  (
    "Arenbergkasteel", "university", 50.8647, 4.6782, "arenbergkasteel-park",
    dict(en="Arenberg Castle & Park",nl="Arenbergkasteel & -park",
         fr="Château d'Arenberg & parc",de="Arenberg-Schloss & Park",
         zh="阿倫貝格城堡與公園"),
    dict(en="A neo-Gothic castle set in a beautiful park, now home to KU Leuven engineering faculty.",
         nl="Een neogotisch kasteel in een prachtig park, nu de thuisbasis van de ingenieursfaculteit.",
         fr="Un château néo-gothique dans un beau parc, abritant la faculté d'ingénierie de KU Leuven.",
         de="Ein neugotisches Schloss in einem schönen Park, heute Sitz der Ingenieursfakultät der KU Leuven.",
         zh="新哥德式城堡坐落於美麗公園中，現為魯汶大學工學院所在地。"),
  ),
  (
    "Anatomisch Theater", "university", 50.8785, 4.7028, "anatomisch-theater",
    dict(en="Anatomical Theatre",nl="Anatomisch Theater",fr="Théâtre anatomique",
         de="Anatomisches Theater",zh="維薩里解剖劇場"),
    dict(en="A historic anatomical theatre once used by Andreas Vesalius, pioneer of modern anatomy.",
         nl="Een historisch anatomisch theater dat ooit gebruikt werd door Andreas Vesalius.",
         fr="Théâtre anatomique historique utilisé par André Vésale, pionnier de l'anatomie moderne.",
         de="Ein historisches anatomisches Theater, das einst von Andreas Vesalius genutzt wurde.",
         zh="歷史性解剖劇場，曾為現代解剖學先驅安德雷亞斯·維薩里使用。"),
  ),
  (
    "Hollands College", "university", 50.8788, 4.7016, "groepen/hollands-college",
    dict(en="Hollands College",nl="Hollands College",fr="Collège Hollandais",
         de="Hollands College",zh="荷蘭學院"),
    dict(en="A 17th-century college with a magnificent Renaissance courtyard, now an events venue.",
         nl="Een 17e-eeuws college met een prachtige renaissance-binnenplaats, nu een evenementenlocatie.",
         fr="Un collège du XVIIe siècle avec une magnifique cour renaissance, maintenant salle d'événements.",
         de="Ein College aus dem 17. Jahrhundert mit einem prächtigen Renaissance-Innenhof.",
         zh="17世紀學院，設有華美的文藝復興庭院，現作為活動場地使用。"),
  ),
  (
    "Janseniuspark", "green", 50.8841, 4.6940, "janseniuspark",
    dict(en="Janseniuspark",nl="Janseniuspark",fr="Parc Jansénius",
         de="Janseniuspark",zh="揚森尼烏斯公園"),
    dict(en="A peaceful green park near the city centre, perfect for a rest between sightseeing.",
         nl="Een rustig groen park dicht bij het stadscentrum, ideaal voor een pauze tijdens het toeren.",
         fr="Un parc verdoyant et tranquille près du centre-ville, idéal pour une pause entre visites.",
         de="Ein ruhiger Grünpark nahe dem Stadtzentrum, ideal für eine Pause beim Sightseeing.",
         zh="靠近市中心的寧靜綠色公園，是觀光途中休憩的理想場所。"),
  ),
  (
    "Museum voor Dierkunde", "museum", 50.8765, 4.6955, "museum-voor-dierkunde",
    dict(en="Museum of Zoology",nl="Museum voor Dierkunde",fr="Musée de zoologie",
         de="Zoologisches Museum",zh="動物學博物館"),
    dict(en="A natural history museum of zoology housed within KU Leuven, free and fascinating.",
         nl="Een natuurhistorisch museum voor dierkunde op de campus van KU Leuven, gratis toegang.",
         fr="Un musée d'histoire naturelle et de zoologie sur le campus de KU Leuven, entrée gratuite.",
         de="Ein naturhistorisches Museum für Zoologie auf dem Campus der KU Leuven, kostenloser Eintritt.",
         zh="魯汶大學校園內的動物學自然歷史博物館，免費參觀，引人入勝。"),
  ),
]

VL_BASE = "https://www.visitleuven.be"

# ─────────────────────────────────────────────────────────────
# 1. 景點
# ─────────────────────────────────────────────────────────────
def get_attractions() -> list[dict]:
    print("  [ATT] 載入多語系景點資料...")
    out = []
    for i, row in enumerate(ATTRACTIONS_ML):
        name, cat, lat, lon, slug, names_d, descs_d = row
        out.append({
            "attraction_id": f"ATT_{i+1:03d}",
            "name":          name,
            "category":      cat,
            "latitude":      lat,
            "longitude":     lon,
            "url":           f"{VL_BASE}/{slug}",
            "description":   descs_d["en"],   # Excel 用英文
            "names":         names_d,          # GeoJSON 多語
            "descriptions":  descs_d,          # GeoJSON 多語
        })
    print(f"     → {len(out)} 個景點（含 EN/NL/FR/DE/ZH 翻譯）")
    return out


# ─────────────────────────────────────────────────────────────
# 2. Google Places
# ─────────────────────────────────────────────────────────────
NEARBY_URL  = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
DETAILS_URL = "https://maps.googleapis.com/maps/api/place/details/json"
DETAIL_FIELDS = ",".join([
    "place_id","name","formatted_address","geometry",
    "opening_hours","price_level","rating","user_ratings_total",
    "reviews","website","types","business_status",
])
PRICE_MAP = {0:"free",1:"€",2:"€€",3:"€€€",4:"€€€€"}


def gp_nearby(lat, lon, radius, ptype, key):
    results, params = [], {
        "location":f"{lat},{lon}","radius":radius,
        "type":ptype,"key":key,"language":"en",
    }
    while True:
        r = requests.get(NEARBY_URL, params=params, timeout=20).json()
        if r.get("status") not in ("OK","ZERO_RESULTS"):
            print(f"     ⚠ {r.get('status')}: {r.get('error_message','')}")
            break
        results.extend(r.get("results",[]))
        tok = r.get("next_page_token")
        if not tok: break
        time.sleep(2.2); params={"pagetoken":tok,"key":key}
    return results


def gp_details(pid, key):
    r = requests.get(DETAILS_URL, params={
        "place_id":pid,"fields":DETAIL_FIELDS,
        "key":key,"language":"en","reviews_no_translations":"true",
    }, timeout=20).json()
    return r.get("result",{})


def infer_reviews(reviews):
    text = " ".join(rv.get("text","") for rv in reviews).lower()
    words = re.findall(r'\w+', text)
    def hit(kws):
        return any(kw.replace(" ","_") in text or
                   all(w in words for w in kw.split()) for kw in kws)
    wifi   = "yes" if hit(WIFI_KW)   else "unknown"
    socket = "yes" if hit(SOCKET_KW) else "unknown"
    quite  = "yes" if hit(QUIET_KW)  else "unknown"
    if any(n in text for n in ("no wifi","no wi-fi","without wifi","pas de wifi")):
        wifi="no"
    if any(n in text for n in ("no socket","no outlet","no plug")):
        socket="no"
    return {"wifi":wifi,"socket":socket,"quite":quite,"review_count":len(reviews)}


def fetch_google_cafes(lat, lon, radius, key):
    print("  [GP] Nearby Search...")
    raw = {}
    for pt in ("cafe","bar"):
        res = gp_nearby(lat, lon, radius, pt, key)
        print(f"     {pt}: {len(res)} 筆")
        for r in res:
            pid = r.get("place_id","")
            if pid and pid not in raw: raw[pid]=r

    print(f"  [GP] {len(raw)} 地點，抓詳細資料...")
    cafes = []
    for idx,(pid,basic) in enumerate(raw.items(),1):
        print(f"     [{idx:3d}/{len(raw)}] {basic.get('name','?')[:40]}", end="\r")
        det = gp_details(pid, key); time.sleep(0.05)

        if det.get("business_status")=="CLOSED_PERMANENTLY": continue

        name  = det.get("name") or basic.get("name","Unknown")
        addr  = det.get("formatted_address") or basic.get("vicinity","")
        geo   = det.get("geometry",{}).get("location",
                basic.get("geometry",{}).get("location",{}))
        plat, plon = geo.get("lat",""), geo.get("lng","")

        types = det.get("types", basic.get("types",[]))
        cafe_type = ("bar" if "bar" in types and "cafe" not in types
                     else "coworking" if any(t in name.lower() for t in ("cowork","work")) else "cafe")

        hours  = det.get("opening_hours",{})
        wk     = hours.get("weekday_text",[])
        open_hr= " | ".join(wk) if wk else "N/A"
        open_now = hours.get("open_now")

        reviews  = det.get("reviews",[])
        inferred = infer_reviews(reviews)

        cafes.append({
            "cafe_id":       f"GP_{pid[:14]}",
            "place_id":      pid,
            "name":          name,
            "cafe_type":     cafe_type,
            "latitude":      plat,
            "longitude":     plon,
            "address":       addr,
            "open_hr":       open_hr,
            "open_now":      "yes" if open_now else ("no" if open_now is False else "unknown"),
            "wifi":          inferred["wifi"],
            "socket":        inferred["socket"],
            "quite":         inferred["quite"],
            "wifi_note":     "inferred from Google reviews" if inferred["wifi"]=="yes" else "",
            "price_avg":     PRICE_MAP.get(det.get("price_level"),"unknown"),
            "rating":        det.get("rating",""),
            "ratings_total": det.get("user_ratings_total",""),
            "review_count":  inferred["review_count"],
            "parking":       "unknown",
            "maps_url":      f"https://www.google.com/maps/place/?q=place_id:{pid}",
            "website":       det.get("website",""),
            "source":        "Google Places",
        })
    print(f"\n  ✅ Google Places：{len(cafes)} 間（排除永久關閉）")
    return cafes


# ─────────────────────────────────────────────────────────────
# 3. OSM WiFi 補充
# ─────────────────────────────────────────────────────────────
def fetch_osm_wifi(lat, lon, radius):
    url = "https://overpass-api.de/api/interpreter"
    q = f"""[out:json][timeout:90];
    (node["amenity"="cafe"]["internet_access"](around:{radius},{lat},{lon});
     node["amenity"="cafe"]["socket"](around:{radius},{lat},{lon});
     node["coworking"="yes"](around:{radius},{lat},{lon});
     node["laptop_friendly"="yes"](around:{radius},{lat},{lon}););out body;"""
    try:
        resp = requests.post(url, data={"data":q}, timeout=120); resp.raise_for_status()
        result = {}
        WYES = {"wlan","yes","free","customers"}
        for e in resp.json().get("elements",[]):
            tags = e.get("tags",{})
            name = (tags.get("name") or "").lower().strip()
            if not name: continue
            ia = tags.get("internet_access","").lower()
            wifi = "yes" if ia in WYES else ("no" if ia=="no" else "unknown")
            if wifi=="unknown":
                for k in ("wifi","free_wifi"):
                    if tags.get(k,"").lower() in ("yes","free"): wifi="yes"; break
            socket = "unknown"
            for k in ("socket","power_supply","charging","outlets"):
                v = tags.get(k,"").lower()
                if v in ("yes","free","limited"): socket="yes"; break
                elif v=="no": socket="no"; break
            if tags.get("coworking")=="yes" or tags.get("laptop_friendly")=="yes": socket="yes"
            result[name] = {"wifi":wifi,"socket":socket}
        print(f"  [OSM] WiFi tag 補充：{len(result)} 筆")
        return result
    except Exception as e:
        print(f"  [OSM] 失敗（略過）：{e}"); return {}


def merge_osm(cafes, osm):
    for c in cafes:
        key = c["name"].lower().strip()
        hit = osm.get(key)
        if not hit:
            for k,v in osm.items():
                if k[:10]==key[:10]: hit=v; break
        if hit:
            if c["wifi"]=="unknown"   and hit["wifi"]  !="unknown": c["wifi"]  =hit["wifi"];   c["wifi_note"]="from OSM tag"
            if c["socket"]=="unknown" and hit["socket"]!="unknown": c["socket"]=hit["socket"]
    return cafes


# ─────────────────────────────────────────────────────────────
# 4. 空間配對
# ─────────────────────────────────────────────────────────────
def haversine(la1,lo1,la2,lo2):
    R=6371000
    try:
        p1,p2=math.radians(float(la1)),math.radians(float(la2))
        dp=math.radians(float(la2)-float(la1)); dl=math.radians(float(lo2)-float(lo1))
        a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
        return round(R*2*math.atan2(math.sqrt(a),math.sqrt(1-a)),1)
    except: return float("inf")

WR={"yes":0,"unknown":1,"no":2}

def spatial_join(attractions, cafes, max_dist, top_n):
    pairs=[]
    for att in attractions:
        nearby=[{**c,"distance_m":haversine(att["latitude"],att["longitude"],c["latitude"],c["longitude"])}
                for c in cafes if haversine(att["latitude"],att["longitude"],c["latitude"],c["longitude"])<=max_dist]
        nearby.sort(key=lambda c:(WR.get(c["wifi"],1),WR.get(c["socket"],1),c["distance_m"]))
        for rank,cafe in enumerate(nearby[:top_n],1):
            pairs.append({
                "attraction_id":att["attraction_id"],"attraction_name":att["name"],
                "att_category":att["category"],"att_lat":att["latitude"],"att_lon":att["longitude"],
                "att_url":att["url"],"cafe_rank":rank,
                "cafe_id":cafe["cafe_id"],"cafe_name":cafe["name"],"cafe_type":cafe["cafe_type"],
                "distance_m":cafe["distance_m"],"rating":cafe.get("rating",""),
                "ratings_total":cafe.get("ratings_total",""),
                "wifi":cafe["wifi"],"socket":cafe["socket"],"quite":cafe["quite"],
                "open_hr":cafe["open_hr"],"open_now":cafe.get("open_now","unknown"),
                "price_avg":cafe["price_avg"],"maps_url":cafe["maps_url"],
                "cafe_lat":cafe["latitude"],"cafe_lon":cafe["longitude"],
            })
    print(f"  ✅ 空間配對：{len(pairs)} 組")
    return pairs


# ─────────────────────────────────────────────────────────────
# 5. GeoJSON — 景點含多語系欄位
# ─────────────────────────────────────────────────────────────
def build_geojson(attractions, cafes, pairs):
    features=[]
    att_cafes={}
    for p in pairs:
        att_cafes.setdefault(p["attraction_id"],[]).append({
            "cafe_id":p["cafe_id"],"name":p["cafe_name"],
            "distance_m":p["distance_m"],"wifi":p["wifi"],
            "socket":p["socket"],"rating":p.get("rating",""),
            "maps_url":p["maps_url"],"cafe_type":p["cafe_type"],
            "open_now":p.get("open_now","unknown"),
        })

    for att in attractions:
        features.append({
            "type":"Feature",
            "geometry":{"type":"Point","coordinates":[float(att["longitude"]),float(att["latitude"])]},
            "properties":{
                "feature_type":"attraction",
                "id":att["attraction_id"],
                "name":att["name"],            # canonical (荷語/英語混合)
                "names":att.get("names",{}),   # ← 多語版本
                "category":att["category"],
                "description":att["description"],
                "descriptions":att.get("descriptions",{}),  # ← 多語版本
                "url":att["url"],
                "nearby_cafes":att_cafes.get(att["attraction_id"],[]),
            }
        })

    for c in cafes:
        features.append({
            "type":"Feature",
            "geometry":{"type":"Point","coordinates":[float(c["longitude"]),float(c["latitude"])]},
            "properties":{
                "feature_type":"cafe","id":c["cafe_id"],"name":c["name"],
                "cafe_type":c["cafe_type"],"address":c["address"],
                "open_hr":c["open_hr"],"open_now":c.get("open_now","unknown"),
                "wifi":c["wifi"],"socket":c["socket"],"quite":c["quite"],
                "price_avg":c["price_avg"],"rating":c.get("rating",""),
                "ratings_total":c.get("ratings_total",""),
                "review_count":c.get("review_count",0),
                "maps_url":c["maps_url"],"website":c.get("website",""),
                "source":c.get("source",""),
            }
        })

    return {
        "type":"FeatureCollection",
        "metadata":{
            "generated":time.strftime("%Y-%m-%dT%H:%M:%SZ",time.gmtime()),
            "city":"Leuven, Belgium",
            "attractions":len(attractions),"cafes":len(cafes),"pairs":len(pairs),
            "languages":["en","nl","fr","de","zh"],
            "source":"Google Places API + OSM Overpass",
        },
        "features":features
    }


# ─────────────────────────────────────────────────────────────
# 6. Excel
# ─────────────────────────────────────────────────────────────
def _brd():
    s=Side(style="thin",color="D0D0D0"); return Border(left=s,right=s,top=s,bottom=s)

def _hdr(ws,color=THEME):
    b=_brd()
    for cell in ws[1]:
        cell.font=Font(name="Calibri",bold=True,color="FFFFFF",size=11)
        cell.fill=PatternFill("solid",fgColor=color)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=b
    ws.row_dimensions[1].height=30

def _body(ws,alt="F0F7F0"):
    b,af=_brd(),PatternFill("solid",fgColor=alt)
    for i,row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row),2):
        for cell in row:
            cell.font=Font(name="Calibri",size=10)
            cell.fill=af if i%2==0 else PatternFill()
            cell.alignment=Alignment(vertical="center"); cell.border=b
        ws.row_dimensions[i].height=16

WF={"yes":"C8E6C9","no":"FFCDD2","unknown":"FFF9C4"}
WC={"yes":("1B5E20",True),"no":("B71C1C",False),"unknown":("F57F17",False)}
SF={"yes":"BBDEFB","no":"FFE0B2","unknown":"F3E5F5"}
SC={"yes":("0D47A1",True),"no":("BF360C",False),"unknown":("6A1B9A",False)}

def _colorcol(ws,cl,fills,fonts):
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
        for cell in row:
            if get_column_letter(cell.column)==cl:
                v=str(cell.value or "unknown").lower()
                cell.fill=PatternFill("solid",fgColor=fills.get(v,"FFFFFF"))
                fc,bold=fonts.get(v,("333333",False))
                cell.font=Font(name="Calibri",size=10,color=fc,bold=bold)

def _linkcol(ws,cl):
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
        for cell in row:
            if get_column_letter(cell.column)==cl and cell.value:
                cell.hyperlink=str(cell.value)
                cell.font=Font(name="Calibri",size=10,color="1155CC",underline="single")

def _widths(ws,d):
    h=[c.value for c in ws[1]]
    for i,v in enumerate(h,1):
        ws.column_dimensions[get_column_letter(i)].width=d.get(v,14)


def save_excel(attractions,cafes,pairs,filepath):
    # flatten names/descriptions dicts for Excel
    rows_att=[]
    for a in attractions:
        row={k:v for k,v in a.items() if k not in ("names","descriptions")}
        for l in ("en","nl","fr","de","zh"):
            row[f"name_{l}"]=a.get("names",{}).get(l,"")
            row[f"desc_{l}"]=a.get("descriptions",{}).get(l,"")
        rows_att.append(row)

    df_att=pd.DataFrame(rows_att); df_att.insert(0,"No.",range(1,len(df_att)+1))
    df_cafe=pd.DataFrame(cafes);   df_cafe.insert(0,"No.",range(1,len(df_cafe)+1))
    df_pairs=pd.DataFrame(pairs);  df_pairs.insert(0,"No.",range(1,len(df_pairs)+1))

    with pd.ExcelWriter(filepath,engine="openpyxl") as w:
        df_att.to_excel(w,  sheet_name="🏛 Attractions",index=False)
        df_cafe.to_excel(w, sheet_name="☕ Cafes",      index=False)
        df_pairs.to_excel(w,sheet_name="📍 Pairs",      index=False)
        pd.DataFrame().to_excel(w,sheet_name="📊 Summary",index=False)

    wb=load_workbook(filepath)
    ws1=wb["🏛 Attractions"]; _hdr(ws1,THEME); _body(ws1,"EAF3DE")
    ws1.freeze_panes="B2"; ws1.auto_filter.ref=ws1.dimensions
    _widths(ws1,{"No.":4,"attraction_id":14,"name":22,"category":12,
                 "latitude":11,"longitude":11,"url":20,"description":36,
                 "name_en":22,"name_nl":22,"name_fr":22,"name_de":22,"name_zh":16,
                 "desc_en":36,"desc_nl":36,"desc_fr":36,"desc_de":36,"desc_zh":30})
    h1=[c.value for c in ws1[1]]
    if "url" in h1: _linkcol(ws1,get_column_letter(h1.index("url")+1))

    ws2=wb["☕ Cafes"]; _hdr(ws2,"2E7D32"); _body(ws2,"F1F8E9")
    ws2.freeze_panes="C2"; ws2.auto_filter.ref=ws2.dimensions
    _widths(ws2,{"No.":4,"cafe_id":18,"name":26,"cafe_type":12,"latitude":11,"longitude":11,
                 "address":32,"open_hr":38,"open_now":10,"wifi":9,"socket":9,"quite":9,
                 "price_avg":10,"rating":10,"ratings_total":10,"maps_url":14,"website":20,"source":16})
    h2=[c.value for c in ws2[1]]
    for col,f,c in [("wifi",WF,WC),("socket",SF,SC)]:
        if col in h2: _colorcol(ws2,get_column_letter(h2.index(col)+1),f,c)
    for col in ("maps_url","website"):
        if col in h2: _linkcol(ws2,get_column_letter(h2.index(col)+1))
    if "rating" in h2:
        rc=h2.index("rating")+1
        for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row):
            cell=row[rc-1]
            try:
                v=float(cell.value or 0)
                cell.value=f"{'★'*int(v)}{'☆'*(5-int(v))} {v:.1f}"
                cell.font=Font(name="Calibri",size=10,
                  color="1B5E20" if v>=4.0 else ("F57F17" if v>=3.0 else "B71C1C"))
            except: pass

    ws3=wb["📍 Pairs"]; _hdr(ws3,"1565C0"); _body(ws3,"E3F2FD")
    ws3.freeze_panes="C2"; ws3.auto_filter.ref=ws3.dimensions
    _widths(ws3,{"No.":4,"attraction_id":14,"attraction_name":26,"att_category":12,
                 "cafe_rank":8,"cafe_id":18,"cafe_name":26,"distance_m":12,
                 "rating":12,"wifi":9,"socket":9,"open_hr":36,"price_avg":10,"maps_url":14})
    h3=[c.value for c in ws3[1]]
    for col,f,c in [("wifi",WF,WC),("socket",SF,SC)]:
        if col in h3: _colorcol(ws3,get_column_letter(h3.index(col)+1),f,c)
    if "maps_url" in h3: _linkcol(ws3,get_column_letter(h3.index("maps_url")+1))

    ws4=wb["📊 Summary"]
    ws4["A1"]="📊 Leuven Tourism × Cafe v3 — Summary"
    ws4["A1"].font=Font(name="Calibri",bold=True,size=15,color=THEME)
    ws4.merge_cells("A1:C1"); ws4.row_dimensions[1].height=32
    wyes=sum(1 for c in cafes if c["wifi"]=="yes")
    syes=sum(1 for c in cafes if c["socket"]=="yes")
    both=sum(1 for c in cafes if c["wifi"]=="yes" and c["socket"]=="yes")
    rows=[("","Item","Count"),
          ("🏛","Total attractions",len(attractions)),
          ("🌐","Languages supported","EN / NL / FR / DE / 中文"),
          ("☕","Total cafes (Google Places)",len(cafes)),
          ("","WiFi = Yes",wyes),("","Socket = Yes",syes),("","Both = Yes",both),
          ("📍","Total pairs",len(pairs))]
    hfill=PatternFill("solid",fgColor=THEME)
    for r,(ic,item,val) in enumerate(rows,3):
        ws4.cell(r,1,ic); ws4.cell(r,2,item); ws4.cell(r,3,val)
        if r==3:
            for c in (1,2,3):
                ws4.cell(r,c).fill=hfill
                ws4.cell(r,c).font=Font(name="Calibri",bold=True,color="FFFFFF",size=11)
        else:
            ws4.cell(r,2).font=Font(name="Calibri",size=11,bold=bool(ic))
            ws4.cell(r,3).font=Font(name="Calibri",size=11,bold=True)
        for c in (1,2,3):
            ws4.cell(r,c).alignment=Alignment(horizontal="left" if c<3 else "center",vertical="center")
    ws4.column_dimensions["A"].width=5
    ws4.column_dimensions["B"].width=34
    ws4.column_dimensions["C"].width=22
    wb.save(filepath); print(f"  💾 Excel: {filepath}")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    p=argparse.ArgumentParser(description="Leuven Scraper v3 — multilingual")
    p.add_argument("--google-key",required=True)
    p.add_argument("--radius",   type=int,default=2500)
    p.add_argument("--pair-dist",type=int,default=500)
    p.add_argument("--top-n",    type=int,default=5)
    p.add_argument("--no-osm",   action="store_true")
    args=p.parse_args()

    os.makedirs(OUTPUT_DIR,exist_ok=True)
    print("\n🇧🇪  Leuven Scraper v3 (multilingual GeoJSON)")
    print(f"   半徑:{args.radius}m  配對:{args.pair_dist}m  Top-N:{args.top_n}")

    print("\n[1/5] 景點..."); attractions=get_attractions()
    print("\n[2/5] Google Places..."); cafes=fetch_google_cafes(LEUVEN_LAT,LEUVEN_LON,args.radius,args.google_key)
    if not args.no_osm:
        print("\n[3/5] OSM WiFi補充...")
        cafes=merge_osm(cafes,fetch_osm_wifi(LEUVEN_LAT,LEUVEN_LON,args.radius))
    else:
        print("\n[3/5] OSM 跳過")
    print("\n[4/5] 空間配對..."); pairs=spatial_join(attractions,cafes,args.pair_dist,args.top_n)
    print("\n[5/5] 輸出...")
    save_excel(attractions,cafes,pairs,os.path.join(OUTPUT_DIR,EXCEL_FILE))
    gj=build_geojson(attractions,cafes,pairs)
    jp=os.path.join(OUTPUT_DIR,JSON_FILE)
    with open(jp,"w",encoding="utf-8") as f: json.dump(gj,f,ensure_ascii=False,indent=2)
    print(f"  💾 GeoJSON: {jp}")

    wyes=sum(1 for c in cafes if c["wifi"]=="yes")
    print(f"\n{'='*50}\n  ✅ 完成！景點:{len(attractions)}  咖啡:{len(cafes)}(WiFi✓{wyes})  配對:{len(pairs)}\n  GeoJSON 含 EN/NL/FR/DE/ZH 多語翻譯\n{'='*50}\n")

if __name__=="__main__": main()
